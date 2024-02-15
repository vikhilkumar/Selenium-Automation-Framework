import csv
import inspect
import logging

import softest
from openpyxl import load_workbook,workbook


class Utils(softest.TestCase):
    def assertListItemText(self,stops_list,stops_value):
        for stop in stops_list:
            print("The stop is: " + stop.text)
            self.soft_assert(self.assertEqual,stop.text,stops_value)
            if stop.text==stops_value:
                print("test pass")
            else:
                print("test failed")     #difference between utils and base_driver folder is if code contains driver then we have to use it in base_driver folder else in utils.
        # self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # Set class/method name from where it's called
        logger_name = inspect.stack()[1][3]

        # Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # Create file handler and set the log level
        fh = logging.FileHandler('../utilities/automation.log',mode='w')

        # Create formatter - specify how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s',datefmt='%m %d %Y %I%M%S %p')

        # Add formatter to the file handler
        fh.setFormatter(formatter)

        # Add file handler to the logger
        logger.addHandler(fh)

        return logger

    def read_data_from_excel(file_name,sheet):
        datalist=[]
        wb=load_workbook(filename=file_name)
        sh=wb['Sheet1']
        row_ct=sh.max_row
        col_ct=sh.max_column

        for i in range(2,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        #create an empty list
        datalist=[]
        #open csv file
        csvdata=open(filename,'r')
        #create csv reader
        reader=csv.reader(csvdata)
        #skip header
        next(reader)
        #Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist

